import mysql.connector
import openpyxl

path = r"D:\offc\IS_Billing\IS_Billing\Data\Process_Invoice.xlsx"

wb=openpyxl.load_workbook(path)

sheet_obj = wb.active

max_col = sheet_obj.max_column
print(max_col)
max_row = sheet_obj.max_row
#print(max_row)
lst=[]
for i in range(2, max_row + 1):
   #for j in range(0,4)
   #for j in range(2,max_col+1):
##      
   VN=str(sheet_obj.cell(row=i,column=1).value)
   VC=str(sheet_obj.cell(row=i,column=2).value)
   PO=str(sheet_obj.cell(row=i,column=3).value)
   INV=str(sheet_obj.cell(row=i,column=4).value)
   TA=str(sheet_obj.cell(row=i,column=5).value)
   ID=str(sheet_obj.cell(row=i,column=6).value)
   IPD=str(sheet_obj.cell(row=i,column=7).value)
   LOC=str(sheet_obj.cell(row=i,column=8).value)
   BPN=str(sheet_obj.cell(row=i,column=9).value)
   AS=str(sheet_obj.cell(row=i,column=10).value)
   FS=str(sheet_obj.cell(row=i,column=11).value)
   #print(VN,VC,PO,INV,TA,ID,IPD,LOC,BPN,AS,FS)

   myconn = mysql.connector.connect(host = "localhost", user = "root",passwd = "",database="hero_db")
   cur=myconn.cursor()
   cur.execute("Insert into istable values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(VN,VC,PO,INV,TA,ID,IPD,LOC,BPN,AS,FS))
   myconn.commit()
myconn.close()
print("data inserted sucessfully")
   
   
   
   

   
